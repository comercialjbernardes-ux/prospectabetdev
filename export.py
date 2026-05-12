"""
export.py — Exportação de dados (CSV / XLSX)
=============================================
Lógica isolada para o endpoint `/api/exportar`. Recebe a lista de registros
já filtrada e retorna a resposta Flask pronta (Response com mimetype).

API pública:
    exportar(dados, formato="csv") -> flask.Response

Comportamento:
- formato="xlsx" tenta `openpyxl`; se não instalado, faz fallback para CSV.
- CSV é UTF-8 com BOM para abrir bem no Excel.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Iterable

from flask import Response


COLUNAS_PADRAO: tuple[str, ...] = (
    "marca", "razao_social", "cnpj", "url", "email_contato", "status",
    "url_afiliados", "status_afiliados",
    "regime_tributario", "porte_empresa", "situacao_cadastral", "capital_social",
    "natureza_juridica", "data_abertura", "logradouro", "numero", "complemento",
    "bairro", "municipio", "uf", "cep", "pais",
    "fonte_regime", "confiabilidade_dado", "data_coleta", "observacao",
)


def _csv_celula(v: object) -> str:
    s = str(v or "")
    if any(c in s for c in (",", '"', "\n")):
        return '"' + s.replace('"', '""') + '"'
    return s


def _resp_csv(dados: Iterable[dict], colunas: tuple[str, ...], nome_arquivo: str) -> Response:
    linhas = [",".join(colunas)]
    for r in dados:
        linhas.append(",".join(_csv_celula(r.get(c, "")) for c in colunas))
    # BOM para compatibilidade com Excel
    conteudo = "﻿" + "\r\n".join(linhas)
    return Response(
        conteudo,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}.csv"'},
    )


def _resp_xlsx(dados: Iterable[dict], colunas: tuple[str, ...], nome_arquivo: str) -> Response | None:
    """Retorna Response XLSX ou None se openpyxl indisponível."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bets"
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="E2E8F0", bold=True)
    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")
    for row_idx, r in enumerate(dados, 2):
        for col_idx, col in enumerate(colunas, 1):
            ws.cell(row=row_idx, column=col_idx, value=r.get(col, ""))
    # Autofit
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}.xlsx"'},
    )


def exportar(
    dados: list[dict],
    formato: str = "csv",
    colunas: tuple[str, ...] = COLUNAS_PADRAO,
) -> Response:
    """Exporta `dados` no `formato` desejado. XLSX faz fallback para CSV se openpyxl ausente."""
    fmt = (formato or "csv").lower()
    nome_arquivo = f"bets_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    if fmt == "xlsx":
        resp = _resp_xlsx(dados, colunas, nome_arquivo)
        if resp is not None:
            return resp
    return _resp_csv(dados, colunas, nome_arquivo)
