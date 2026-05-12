"""
app.py — Dashboard Flask para visualização de bets regulamentadas
================================================================
Uso:
    python _start_server.py      (Waitress, produção)
    python app.py                (Flask dev, porta 5000)
    Acesse: http://127.0.0.1:5002

Edição manual:
    POST /api/editar  →  overrides.json  →  audit_log.jsonl

Paginação server-side (opcional):
    GET /api/dados?pagina=1&limite=50   → retorna envelope {dados, total, ...}
    GET /api/dados                      → retorna lista completa (retrocompat.)
"""

import csv
import io
import json
import os
import re
import shutil
import time
from datetime import datetime
from pathlib import Path
from threading import Lock, RLock
from urllib.parse import urlparse

from flask import Flask, Response, jsonify, render_template, request
import json_store
from logging_config import get_logger

logger = get_logger(__name__)

import url_health
import csv_sync

try:
    import afiliados_health
    _AFILIADOS_HEALTH_DISPONIVEL = True
except ImportError:
    _AFILIADOS_HEALTH_DISPONIVEL = False

try:
    import reclame_aqui_health as _ra_health
    _RA_HEALTH_DISPONIVEL = True
except ImportError:
    _ra_health = None  # type: ignore
    _RA_HEALTH_DISPONIVEL = False

# Playwright disponível?
try:
    from playwright.sync_api import sync_playwright as _pw  # noqa: F401
    _PLAYWRIGHT_DISPONIVEL = True
except ImportError:
    _PLAYWRIGHT_DISPONIVEL = False

app = Flask(__name__)

ARQUIVO_JSON      = Path("dados/bets_enriquecidas.json")
ARQUIVO_CSV       = Path("bets_com_emails.csv")
ARQUIVO_OVERRIDES = Path("dados/overrides.json")
ARQUIVO_AUDIT     = Path("dados/audit_log.jsonl")

# Audit log: limite de linhas antes de truncar
_AUDIT_MAX_LINHAS    = 50_000
_AUDIT_TRUNCAR_PARA  = 40_000   # linhas mais recentes mantidas ao truncar

# Cache de stats
_STATS_CACHE_TTL   = 5.0        # segundos

# Validação de campos
_CAMPO_MAX_TAMANHO = 2_000       # caracteres
_CNPJ_SODIGITOS_RE = re.compile(r"^\d{14}$")

# Campos editáveis via /api/editar
CAMPOS_EDITAVEIS = {
    "email_contato", "url", "marca", "razao_social",
    "cnpj", "uf", "municipio", "url_afiliados", "observacao",
}

# ---------------------------------------------------------------------------
# Estado global (thread-safe)
# ---------------------------------------------------------------------------

_dados: list[dict] = []
_dados_lock        = RLock()        # protege leitura/escrita de _dados

_overrides: dict[str, dict] = {}
_lock_overrides = Lock()

_AUDIT_LOCK = Lock()

_stats_cache: dict = {"data": None, "ts": 0.0}
_stats_lock  = Lock()


def _snapshot_dados() -> list[dict]:
    """Retorna cópia thread-safe de _dados (evita race condition com workers)."""
    with _dados_lock:
        return list(_dados)


def _invalidar_cache_stats() -> None:
    with _stats_lock:
        _stats_cache["ts"] = 0.0


# ---------------------------------------------------------------------------
# Validações de input
# ---------------------------------------------------------------------------

def _validar_cnpj_formato(cnpj: str) -> bool:
    """Aceita CNPJ com ou sem máscara; valida que tenha 14 dígitos."""
    return bool(_CNPJ_SODIGITOS_RE.match(re.sub(r"\D", "", cnpj)))


def _validar_url_segura(url: str) -> bool:
    """Garante que a URL usa http/https e tem netloc não-vazio."""
    try:
        p = urlparse(url)
        return p.scheme in ("http", "https") and bool(p.netloc)
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Audit log — dados/audit_log.jsonl (append-only, JSON Lines)
# ---------------------------------------------------------------------------

def _rotacionar_audit_log() -> None:
    """Trunca audit_log.jsonl para as N linhas mais recentes se exceder o limite."""
    if not ARQUIVO_AUDIT.exists():
        return
    with _AUDIT_LOCK:
        try:
            with open(ARQUIVO_AUDIT, encoding="utf-8") as fh:
                linhas = fh.readlines()
            if len(linhas) <= _AUDIT_MAX_LINHAS:
                return
            linhas = linhas[-_AUDIT_TRUNCAR_PARA:]
            tmp = ARQUIVO_AUDIT.with_suffix(".jsonl.tmp")
            tmp.write_text("".join(linhas), encoding="utf-8")
            tmp.replace(ARQUIVO_AUDIT)
            logger.info(f"audit_log truncado: mantidas {len(linhas)} linhas mais recentes")
        except OSError:
            pass


def _registrar_auditoria(
    acao: str, cnpj: str, campo: str,
    valor_anterior, valor_novo, ip: str = "",
) -> None:
    """Acrescenta uma linha JSON ao audit log (append-only, thread-safe)."""
    entrada = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "acao": acao, "cnpj": cnpj, "campo": campo,
        "valor_anterior": valor_anterior, "valor_novo": valor_novo, "ip": ip,
    }
    ARQUIVO_AUDIT.parent.mkdir(parents=True, exist_ok=True)
    with _AUDIT_LOCK:
        with open(ARQUIVO_AUDIT, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(entrada, ensure_ascii=False) + "\n")
    logger.info("audit", extra=entrada)
    # Rotaciona assincronamente (checagem rápida — só rotaciona se necessário)
    _rotacionar_audit_log()


# ---------------------------------------------------------------------------
# Overrides (edições manuais) — persistidos em dados/overrides.json
# ---------------------------------------------------------------------------

SCHEMA_OVERRIDES_VERSION = 1


def _carregar_overrides() -> dict[str, dict]:
    if not ARQUIVO_OVERRIDES.exists():
        return {}
    try:
        data = json.loads(ARQUIVO_OVERRIDES.read_text("utf-8"))
    except json.JSONDecodeError:
        backup = ARQUIVO_OVERRIDES.with_suffix(f".json.corrupto_{int(time.time())}")
        try:
            shutil.copy2(ARQUIVO_OVERRIDES, backup)
        except Exception:
            pass
        return {}
    except OSError:
        return {}
    versao = data.get("_schema_version", 1)
    if versao < SCHEMA_OVERRIDES_VERSION:
        data = _migrar_overrides(data, de=versao, para=SCHEMA_OVERRIDES_VERSION)
    return {k: v for k, v in data.items() if not k.startswith("_schema")}


def _salvar_overrides(overrides: dict[str, dict]) -> None:
    ARQUIVO_OVERRIDES.parent.mkdir(parents=True, exist_ok=True)
    tmp = ARQUIVO_OVERRIDES.with_suffix(".json.tmp")
    payload = {
        "_schema_version": SCHEMA_OVERRIDES_VERSION,
        "_salvo_em": datetime.now().isoformat(timespec="seconds"),
        **overrides,
    }
    try:
        tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), "utf-8")
        if ARQUIVO_OVERRIDES.exists():
            shutil.copy2(ARQUIVO_OVERRIDES, ARQUIVO_OVERRIDES.with_suffix(".json.bak"))
        tmp.replace(ARQUIVO_OVERRIDES)
    except Exception:
        tmp.unlink(missing_ok=True)
        raise


def _migrar_overrides(data: dict, de: int, para: int) -> dict:
    return data


def _aplicar_overrides(registros: list[dict], overrides: dict[str, dict]) -> None:
    """Mescla overrides nos registros — marca cada registro editado."""
    for r in registros:
        cnpj = (r.get("cnpj") or "").strip()
        if not cnpj or cnpj not in overrides:
            continue
        ov = overrides[cnpj]
        campos_editados = []
        for campo, valor in ov.items():
            if campo.startswith("_") or campo not in CAMPOS_EDITAVEIS:
                continue
            r[campo] = valor
            campos_editados.append(campo)
            if campo == "email_contato":
                if valor:
                    if r.get("status") in (None, "", "nao_encontrado", "erro_conexao",
                                           "bloqueado_robots", "sem_url"):
                        r["status"] = "encontrado_manual"
                else:
                    r["status"] = "nao_encontrado"
            if campo == "url_afiliados":
                if valor:
                    r["status_afiliados"]  = "encontrado_manual"
                    r["_afiliados_status"] = "encontrado_manual"
                    r["_afiliados_display"] = "sim"
                    r["_afiliados_url"]    = valor
                else:
                    r["status_afiliados"] = ""
                    r.pop("_afiliados_status",  None)
                    r.pop("_afiliados_display", None)
                    r.pop("_afiliados_url",     None)
        if campos_editados:
            r["_editado_manualmente"] = True
            r["_campos_editados"]     = campos_editados
            r["_editado_em"]          = ov.get("_edited_at", "")


# ---------------------------------------------------------------------------
# Carregamento de dados
# ---------------------------------------------------------------------------

def _carregar_dados() -> list[dict]:
    dados = json_store.ler(ARQUIVO_JSON, default=None)
    if dados is not None:
        for r in dados:
            try:
                r["capital_social"] = float(r.get("capital_social") or 0)
            except (ValueError, TypeError):
                r["capital_social"] = 0.0
        return dados
    if ARQUIVO_CSV.exists():
        registros = []
        with open(ARQUIVO_CSV, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                registros.append(row)
        return registros
    return []


# Statuses que indicam site INACESSÍVEL (offline/DNS falhou/timeout)
# "bloqueado"  = site ATIVO mas bloqueando bots (403/4xx) → NÃO é inativo
# "erro_http"  = erro HTTP do servidor (5xx) → é inativo
# "redirect"   = redireciona para outro domínio → site ativo
_STATUS_INATIVO = {"erro_http", "erro_conexao", "erro_ssl", "erro_dns", "timeout", "erro"}
_STATUS_ERRO = _STATUS_INATIVO  # alias de compatibilidade

# Afiliados
_STATUS_AFILIADOS_SIM = {"encontrado_completo", "encontrado_url", "encontrado_email"}
_STATUS_AFILIADOS_NAO = {"nao_encontrado", "bloqueado_robots"}


def _display_afiliados(status: str) -> str:
    if status in _STATUS_AFILIADOS_SIM: return "sim"
    if status in _STATUS_AFILIADOS_NAO: return "nao"
    return "nao_encontrado"


def _aplicar_afiliados_health(registros: list[dict]) -> None:
    path = Path("dados/afiliados_health.json")
    if not path.exists():
        for r in registros:
            if r.get("_afiliados_status") != "encontrado_manual":
                if "_afiliados_display" not in r:
                    r["_afiliados_display"] = "nao_encontrado"
        return
    try:
        dados_h = json.loads(path.read_text("utf-8"))
    except Exception:
        return
    for r in registros:
        if r.get("_afiliados_status") == "encontrado_manual":
            continue
        u    = (r.get("url") or "").strip()
        info = dados_h.get(u)
        if not info:
            r["_afiliados_display"] = "nao_encontrado"
            continue
        st = info.get("status", "")
        r["_afiliados_status"]  = st
        r["_afiliados_display"] = _display_afiliados(st)
        r["_afiliados_url"]     = info.get("url_afiliado", "")
        r["_afiliados_ts"]      = info.get("checado_em", "")


def _aplicar_reclame_aqui_health(registros: list[dict]) -> None:
    """Mescla dados do Reclame Aqui (reclame_aqui_health.json) nos registros."""
    if not _RA_HEALTH_DISPONIVEL:
        for r in registros:
            r.setdefault("_ra_status", "desconhecido")
        return
    try:
        dados_h = _ra_health.ler_health()
    except Exception:
        dados_h = {}
    for r in registros:
        slug = _ra_health.slug_para_marca(r.get("marca") or "")
        info = dados_h.get(slug)
        if not info:
            r["_ra_status"] = "desconhecido"
            continue
        r["_ra_status"]      = info.get("status", "desconhecido")
        r["_ra_nota"]        = info.get("nota")
        r["_ra_reclamacoes"] = info.get("total_reclamacoes")
        r["_ra_resolvidas"]  = info.get("percentual_resolvidas")
        r["_ra_reputacao"]   = info.get("reputacao", "")
        r["_ra_ra1000"]      = info.get("ra1000", False)
        r["_ra_url"]         = info.get("url_reclame_aqui", "")
        r["_ra_ts"]          = info.get("checado_em", "")


def _aplicar_url_health(registros: list[dict]) -> None:
    try:
        health = url_health.ler_health()
    except Exception:
        health = {}
    for r in registros:
        u    = (r.get("url") or "").strip()
        info = health.get(u)
        if not info:
            r["_url_health_status"] = "desconhecido"
            r["_url_inativa"]       = False
            continue
        st_raw    = info.get("status", "desconhecido")
        http_code = info.get("http_code", 0)
        # Normaliza status legado: "erro_http" com código 4xx → "bloqueado"
        # (site ATIVO bloqueando bots via 403/4xx — não é inativo)
        if st_raw == "erro_http" and 0 < http_code < 500:
            st = "bloqueado"
        else:
            st = st_raw
        r["_url_health_status"] = st
        r["_url_http_code"]     = http_code
        r["_url_checked_at"]    = info.get("checado_em", "")
        r["_url_latencia_ms"]   = info.get("latencia_ms", 0)
        if info.get("redirecionou"):
            r["_url_redirect_to"] = info.get("url_final", "")
        r["_url_inativa"] = (st in _STATUS_INATIVO) or bool(r.get("_removido_do_csv"))


def recarregar_dados() -> None:
    global _dados, _overrides
    _overrides = _carregar_overrides()
    dados = _carregar_dados()
    _aplicar_overrides(dados, _overrides)
    _aplicar_url_health(dados)
    _aplicar_afiliados_health(dados)
    _aplicar_reclame_aqui_health(dados)
    with _dados_lock:
        _dados = dados
    _invalidar_cache_stats()
    # Salva snapshot diário de stats (para sparklines)
    try:
        import stats_snapshot
        stats_snapshot.registrar_snapshot_se_necessario(dados)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Helpers de filtro (reutilizados por /api/dados paginado e /api/exportar)
# ---------------------------------------------------------------------------

_STATUS_FALHOU = {"erro_conexao", "bloqueado_robots", "sem_url"}


def _aplicar_filtros_query(registros: list[dict], params: dict) -> list[dict]:
    """Filtra registros pelos query params opcionais (mesma lógica do frontend)."""
    marca      = (params.get("marca") or "").lower().strip()
    status     = params.get("status", "")
    afiliados  = params.get("afiliados", "")
    porte      = params.get("porte", "")
    situacao   = params.get("situacao", "")
    uf         = (params.get("uf") or "").upper()
    municipio  = params.get("municipio", "")
    dt_ini     = params.get("data_inicio", "")
    dt_fim     = params.get("data_fim", "")
    saude_url  = params.get("saude_url", "")

    resultado = []
    for r in registros:
        if marca and marca not in (r.get("marca") or "").lower():
            continue
        if status:
            if status == "com_email":
                if not r.get("email_contato"):
                    continue
            elif status == "sem_email":
                if r.get("email_contato"):
                    continue
            elif status == "falhou":
                if r.get("status") not in _STATUS_FALHOU:
                    continue
        if afiliados:
            disp = r.get("_afiliados_display") or "nao_encontrado"
            if afiliados == "com" and disp != "sim":      continue
            if afiliados == "sem" and disp != "nao":      continue
            if afiliados == "nd"  and disp != "nao_encontrado": continue
        if porte    and r.get("porte_empresa")     != porte:    continue
        if situacao and r.get("situacao_cadastral") != situacao: continue
        if uf       and (r.get("uf") or "").upper() != uf:       continue
        if municipio and r.get("municipio") != municipio:        continue
        if dt_ini and (r.get("data_coleta") or "")[:10] < dt_ini: continue
        if dt_fim and (r.get("data_coleta") or "")[:10] > dt_fim: continue
        if saude_url:
            st     = r.get("_url_health_status") or "desconhecido"
            _ONLINE = {"ok", "redirect", "bloqueado"}           # site acessível p/ usuários reais
            _ERROS  = {"erro_http","erro_conexao","erro_ssl","erro_dns","timeout","erro"}
            if saude_url == "online"      and st not in _ONLINE: continue  # todas as acessíveis
            if saude_url == "ok"          and st != "ok":         continue  # 200 direto
            if saude_url == "redirect"    and st != "redirect":   continue  # 30x
            if saude_url == "bloqueado"   and st != "bloqueado":  continue  # 4xx bots
            if saude_url == "erro"        and st not in _ERROS:   continue  # 5xx/offline
            if saude_url == "desconhecido" and st != "desconhecido": continue
        resultado.append(r)
    return resultado


# ---------------------------------------------------------------------------
# Rotas
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/auditoria")
def auditoria():
    """
    Página de auditoria — lê audit_log.jsonl, aplica filtros, pagina e
    enriquece eventos com a marca (mapeada a partir do CNPJ).
    Variáveis Jinja: eventos, total_eventos, pagina_atual, total_paginas,
                     filtros, campos_disponiveis.
    """
    # Filtros vindos da query string
    filtros = {
        "q":          (request.args.get("q") or "").strip(),
        "acao":       (request.args.get("acao") or "").strip().upper(),
        "campo":      (request.args.get("campo") or "").strip(),
        "data_inicio": (request.args.get("data_inicio") or "").strip(),
        "data_fim":    (request.args.get("data_fim") or "").strip(),
    }
    try:
        pagina_atual = max(1, int(request.args.get("page", 1)))
    except ValueError:
        pagina_atual = 1
    POR_PAGINA = 50

    # Mapa CNPJ → marca para enriquecimento (CNPJ do log pode estar mascarado)
    snap = _snapshot_dados()
    def _norm_cnpj(c: str) -> str:
        return re.sub(r"\D", "", c or "")
    cnpj_marca = {_norm_cnpj(r.get("cnpj", "")): r.get("marca", "") for r in snap if r.get("cnpj")}

    # Lê audit_log.jsonl
    eventos_brutos: list[dict] = []
    if ARQUIVO_AUDIT.exists():
        try:
            with open(ARQUIVO_AUDIT, encoding="utf-8") as fh:
                for linha in fh:
                    linha = linha.strip()
                    if not linha:
                        continue
                    try:
                        eventos_brutos.append(json.loads(linha))
                    except json.JSONDecodeError:
                        continue
        except OSError:
            pass
    eventos_brutos.reverse()  # mais recente primeiro

    # Normaliza para o schema do template
    def _normalizar(ev: dict) -> dict:
        ts = ev.get("ts") or ev.get("timestamp", "")
        ts_fmt = ""
        if ts:
            try:
                # 2026-05-02T18:50:01 → 02/05/26 18:50:01
                d = datetime.fromisoformat(ts.replace("Z", ""))
                ts_fmt = d.strftime("%d/%m/%y %H:%M:%S")
            except (ValueError, TypeError):
                ts_fmt = ts
        return {
            "timestamp":     ts,
            "timestamp_fmt": ts_fmt,
            "acao":          (ev.get("acao") or "EDIT").upper(),
            "cnpj":          ev.get("cnpj", ""),
            "marca":         cnpj_marca.get(_norm_cnpj(ev.get("cnpj", "")), ""),
            "campo":         ev.get("campo", ""),
            "valor_anterior": ev.get("valor_anterior") if ev.get("valor_anterior") not in (None, "") else "",
            "valor_novo":     ev.get("valor_novo") if ev.get("valor_novo") not in (None, "") else "",
            "usuario":        ev.get("usuario") or "system",
            "ip":             ev.get("ip", ""),
        }
    eventos_norm = [_normalizar(e) for e in eventos_brutos]

    # Aplica filtros
    def _passa(ev: dict) -> bool:
        if filtros["acao"] and ev["acao"] != filtros["acao"]:
            return False
        if filtros["campo"] and ev["campo"] != filtros["campo"]:
            return False
        if filtros["data_inicio"] and ev["timestamp"][:10] < filtros["data_inicio"]:
            return False
        if filtros["data_fim"] and ev["timestamp"][:10] > filtros["data_fim"]:
            return False
        if filtros["q"]:
            q = filtros["q"].lower()
            blob = " ".join(str(ev.get(k, "")) for k in
                            ("marca", "campo", "usuario", "valor_anterior", "valor_novo")).lower()
            if q not in blob:
                return False
        return True
    eventos_filtrados = [e for e in eventos_norm if _passa(e)]

    total_eventos = len(eventos_filtrados)
    total_paginas = max(1, (total_eventos + POR_PAGINA - 1) // POR_PAGINA)
    pagina_atual = min(pagina_atual, total_paginas)
    inicio = (pagina_atual - 1) * POR_PAGINA
    eventos = eventos_filtrados[inicio:inicio + POR_PAGINA]

    # Lista de campos únicos vistos no log (para o select de filtro)
    campos_disponiveis = sorted({e["campo"] for e in eventos_norm if e["campo"]})

    return render_template(
        "auditoria.html",
        eventos=eventos,
        total_eventos=total_eventos,
        pagina_atual=pagina_atual,
        total_paginas=total_paginas,
        filtros=filtros,
        campos_disponiveis=campos_disponiveis,
    )


@app.route("/api/dados")
def api_dados():
    """
    Retorna registros com merges aplicados.

    Suporte a paginação/filtro server-side (opcional):
        GET /api/dados?pagina=1&limite=50&marca=bet&uf=SP...
    Sem parâmetros: retorna lista completa (retrocompat.).
    """
    snap = _snapshot_dados()
    _aplicar_url_health(snap)
    _aplicar_afiliados_health(snap)

    pagina = request.args.get("pagina", type=int)
    limite = request.args.get("limite", type=int)

    if pagina is not None and limite is not None:
        # Modo paginado com filtros server-side
        filtros = {k: v for k, v in request.args.items()
                   if k not in ("pagina", "limite")}
        filtrado = _aplicar_filtros_query(snap, filtros)
        inicio   = (pagina - 1) * limite
        fim      = inicio + limite
        return jsonify({
            "dados":         filtrado[inicio:fim],
            "total":         len(filtrado),
            "pagina":        pagina,
            "limite":        limite,
            "total_paginas": max(1, (len(filtrado) + limite - 1) // limite),
        })

    return jsonify(snap)


@app.route("/api/stats")
def api_stats():
    """KPI cards com cache de 5 segundos."""
    agora = time.monotonic()
    with _stats_lock:
        if _stats_cache["data"] is not None and (agora - _stats_cache["ts"]) < _STATS_CACHE_TTL:
            return jsonify(_stats_cache["data"])

    snap = _snapshot_dados()
    _aplicar_url_health(snap)
    _aplicar_afiliados_health(snap)

    total      = len(snap)
    com_email  = sum(1 for r in snap if r.get("status") in ("encontrado", "encontrado_js", "encontrado_manual"))
    sem_email  = sum(1 for r in snap if not r.get("email_contato"))
    com_afil   = sum(1 for r in snap if r.get("_afiliados_display") == "sim")
    afil_nao   = sum(1 for r in snap if r.get("_afiliados_display") == "nao")
    afil_nd    = sum(1 for r in snap if r.get("_afiliados_display") == "nao_encontrado")
    editados   = sum(1 for r in snap if r.get("_editado_manualmente"))

    # Última atualização: tenta data_coleta, cai para _enriquecido_em, cai para _adicionado_em
    def _melhor_data(r: dict) -> str:
        return (r.get("data_coleta") or r.get("_enriquecido_em") or r.get("_adicionado_em") or "")

    ultima = max((_melhor_data(r) for r in snap), default="")

    portes     = sorted({r.get("porte_empresa", "")    for r in snap if r.get("porte_empresa")})
    situacoes  = sorted({r.get("situacao_cadastral", "") for r in snap if r.get("situacao_cadastral")})
    ufs        = sorted({r.get("uf", "")               for r in snap if r.get("uf")})

    _URL_ONLINE     = {"ok", "redirect", "bloqueado"}   # site acessível p/ usuários reais
    urls_online     = sum(1 for r in snap if r.get("_url_health_status") in _URL_ONLINE)
    urls_ok         = sum(1 for r in snap if r.get("_url_health_status") == "ok")
    urls_redirect   = sum(1 for r in snap if r.get("_url_health_status") == "redirect")
    urls_bloqueadas = sum(1 for r in snap if r.get("_url_health_status") == "bloqueado")
    urls_inativas   = sum(1 for r in snap if r.get("_url_inativa"))
    urls_desc       = sum(1 for r in snap if r.get("_url_health_status") == "desconhecido")
    # alias de retrocompat. (card usa urls_ativas → agora = online total)
    urls_ativas     = urls_online

    ra_encontradas     = sum(1 for r in snap if r.get("_ra_status") == "encontrado")
    ra_nao_encontradas = sum(1 for r in snap if r.get("_ra_status") == "nao_encontrado")
    ra_pendentes       = sum(1 for r in snap if r.get("_ra_status", "desconhecido") == "desconhecido")

    sync_info = csv_sync.ler_status()

    data = {
        "total": total,
        "com_email": com_email, "sem_email": sem_email,
        "com_afiliados": com_afil, "afiliados_sim": com_afil,
        "afiliados_nao": afil_nao, "afiliados_desconhecido": afil_nd,
        "editados_manualmente": editados, "ultima_atualizacao": ultima,
        "portes": portes, "situacoes": situacoes, "ufs": ufs,
        "urls_ativas": urls_ativas,          # total online (ok + redirect + bloqueado)
        "urls_ok": urls_ok,                  # somente 200 OK direto
        "urls_redirect": urls_redirect,
        "urls_bloqueadas": urls_bloqueadas,
        "urls_inativas": urls_inativas, "urls_desconhecido": urls_desc,
        "ra_encontradas": ra_encontradas,
        "ra_nao_encontradas": ra_nao_encontradas,
        "ra_pendentes": ra_pendentes,
        "playwright_disponivel": _PLAYWRIGHT_DISPONIVEL,
        "csv_sync": {
            "ultimo_sync": sync_info.get("finalizado_em") or sync_info.get("iniciado_em"),
            "sucesso":     sync_info.get("sucesso", False),
            "adicionadas": len(sync_info.get("adicionadas", [])),
            "removidas":   len(sync_info.get("removidas", [])),
            "url_atualizada": len(sync_info.get("url_atualizada", [])),
        },
    }

    with _stats_lock:
        _stats_cache["data"] = data
        _stats_cache["ts"]   = time.monotonic()

    return jsonify(data)


@app.route("/api/municipios/<uf>")
def api_municipios(uf: str):
    municipios = sorted({
        r.get("municipio", "")
        for r in _snapshot_dados()
        if r.get("uf", "").upper() == uf.upper() and r.get("municipio")
    })
    return jsonify(municipios)


@app.route("/api/url-health")
def api_url_health():
    return jsonify(url_health.ler_health())


@app.route("/api/afiliados-health")
def api_afiliados_health():
    path = Path("dados/afiliados_health.json")
    if not path.exists():
        return jsonify({})
    try:
        return jsonify(json.loads(path.read_text("utf-8")))
    except Exception:
        return jsonify({})


@app.route("/api/url-health-worker-status")
def api_url_health_worker_status():
    """Retorna estado do worker de url_health e estatísticas do arquivo."""
    import threading
    worker_vivo = any(
        t.name == "url-health-worker" and t.is_alive()
        for t in threading.enumerate()
    )
    health = url_health.ler_health()
    from datetime import datetime as _dt
    agora = _dt.now()
    idades_min = []
    status_count: dict[str, int] = {}
    for info in health.values():
        st = info.get("status", "desconhecido")
        status_count[st] = status_count.get(st, 0) + 1
        t = info.get("checado_em", "")
        if t:
            try:
                idades_min.append((_dt.now() - _dt.fromisoformat(t)).total_seconds() / 60)
            except Exception:
                pass
    return jsonify({
        "worker_ativo":    worker_vivo,
        "total_urls":      len(health),
        "status_counts":   status_count,
        "idade_min_min":   round(min(idades_min), 1) if idades_min else None,
        "idade_max_min":   round(max(idades_min), 1) if idades_min else None,
        "idade_media_min": round(sum(idades_min) / len(idades_min), 1) if idades_min else None,
    })


@app.route("/api/force-health-recheck", methods=["POST"])
def api_force_health_recheck():
    """
    Apaga os timestamps de `checado_em` do url_health.json, forçando
    o worker a re-validar todas as URLs no próximo tick.
    """
    try:
        health = url_health.ler_health()
        for info in health.values():
            info.pop("checado_em", None)
        import json_store as _js
        _js.salvar(Path("dados/url_health.json"), health)
        return jsonify({"ok": True, "urls_resetadas": len(health)})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


@app.route("/api/reclame-aqui-health")
def api_reclame_aqui_health():
    """Retorna o JSON completo de saúde do Reclame Aqui (chave=slug)."""
    path = Path("dados/reclame_aqui_health.json")
    if not path.exists():
        return jsonify({})
    try:
        return jsonify(json.loads(path.read_text("utf-8")))
    except Exception:
        return jsonify({})


@app.route("/api/csv-sync-status")
def api_csv_sync_status():
    return jsonify(csv_sync.ler_status())


@app.route("/api/csv-sync-agora", methods=["POST"])
def api_csv_sync_agora():
    resultado = csv_sync.sincronizar_uma_vez()
    if resultado.get("sucesso"):
        recarregar_dados()
    return jsonify(resultado)


@app.route("/api/recarregar", methods=["POST"])
def api_recarregar():
    recarregar_dados()
    return jsonify({"ok": True, "total": len(_dados)})


@app.route("/api/audit-log")
def api_audit_log():
    """Retorna as últimas N entradas do audit log (padrão 200, máx 1000)."""
    limite = min(int(request.args.get("limite", 200)), 1000)
    campo  = request.args.get("campo", "")
    cnpj   = request.args.get("cnpj", "")
    if not ARQUIVO_AUDIT.exists():
        return jsonify([])
    try:
        with open(ARQUIVO_AUDIT, encoding="utf-8") as fh:
            linhas = fh.readlines()
    except OSError:
        return jsonify([])
    entradas = []
    for linha in reversed(linhas):
        linha = linha.strip()
        if not linha:
            continue
        try:
            e = json.loads(linha)
        except json.JSONDecodeError:
            continue
        if campo and e.get("campo") != campo:
            continue
        if cnpj and e.get("cnpj") != cnpj:
            continue
        entradas.append(e)
        if len(entradas) >= limite:
            break
    return jsonify(entradas)


@app.route("/api/editar", methods=["POST"])
def api_editar():
    """
    Edita um campo de um registro.

    Body JSON:
        {"cnpj": "12345678000100", "campo": "email_contato", "valor": "x@y.com"}

    valor=null  → reset (volta ao valor base)
    valor=""    → deleção explícita (mascara o valor base)
    """
    payload = request.get_json(silent=True) or {}
    cnpj    = (payload.get("cnpj") or "").strip()
    campo   = (payload.get("campo") or "").strip()
    valor   = payload.get("valor")
    if isinstance(valor, str):
        valor = valor.strip()

    if not cnpj:
        return jsonify({"ok": False, "erro": "CNPJ ausente."}), 400

    # Validação de formato do CNPJ (apenas dígitos, 14 caracteres)
    if not _validar_cnpj_formato(cnpj):
        return jsonify({"ok": False, "erro": "CNPJ deve conter 14 dígitos numéricos."}), 400

    if campo not in CAMPOS_EDITAVEIS:
        return jsonify({
            "ok": False,
            "erro": f"Campo '{campo}' não é editável.",
            "editaveis": sorted(CAMPOS_EDITAVEIS),
        }), 400

    # Validação de tamanho máximo
    if isinstance(valor, str) and len(valor) > _CAMPO_MAX_TAMANHO:
        return jsonify({"ok": False, "erro": f"Valor excede {_CAMPO_MAX_TAMANHO} caracteres."}), 400

    # Validação de email
    if campo == "email_contato" and valor:
        if "@" not in valor or "." not in valor.split("@")[-1]:
            return jsonify({"ok": False, "erro": "Email inválido."}), 400

    # Validação de URL (http/https obrigatório)
    if campo in ("url", "url_afiliados") and valor:
        if not _validar_url_segura(valor):
            return jsonify({"ok": False, "erro": "URL deve começar com http:// ou https://."}), 400

    resetar = valor is None
    deletar = isinstance(valor, str) and valor == ""

    snap = _snapshot_dados()
    registro_atual = next((r for r in snap if (r.get("cnpj") or "").strip() == cnpj), None)
    valor_anterior = registro_atual.get(campo) if registro_atual else None

    with _lock_overrides:
        overrides = _carregar_overrides()
        reg_ov    = overrides.get(cnpj, {})
        if resetar:
            reg_ov.pop(campo, None)
        else:
            reg_ov[campo] = valor
        campos_reais = [k for k in reg_ov if not k.startswith("_")]
        if not campos_reais:
            overrides.pop(cnpj, None)
        else:
            reg_ov["_edited_at"] = datetime.now().isoformat(timespec="seconds")
            overrides[cnpj] = reg_ov
        _salvar_overrides(overrides)

    recarregar_dados()

    acao = "reset" if resetar else ("delete" if deletar else "edit")
    _registrar_auditoria(
        acao=acao, cnpj=cnpj, campo=campo,
        valor_anterior=valor_anterior,
        valor_novo=valor if not resetar else None,
        ip=request.remote_addr or "",
    )

    # Dispara notificação se configurada
    try:
        import notificacoes
        notificacoes.notificar_edicao(cnpj=cnpj, campo=campo,
                                      valor_anterior=valor_anterior, valor_novo=valor)
    except Exception:
        pass

    atualizado = next(
        (r for r in _snapshot_dados() if (r.get("cnpj") or "").strip() == cnpj), None
    )
    return jsonify({"ok": True, "registro": atualizado})


# ---------------------------------------------------------------------------
# FASE 3 — Endpoints adicionais
# ---------------------------------------------------------------------------

@app.route("/api/duplicatas")
def api_duplicatas():
    """
    Retorna emails que aparecem em mais de um registro distinto.
    Útil para identificar dados genéricos (info@, contato@) compartilhados.
    """
    snap = _snapshot_dados()
    contagem: dict[str, list[str]] = {}
    for r in snap:
        email = (r.get("email_contato") or "").strip().lower()
        if not email:
            continue
        marca = r.get("marca") or r.get("cnpj") or "?"
        contagem.setdefault(email, []).append(marca)
    duplicatas = [
        {"email": email, "marcas": marcas, "qtd": len(marcas)}
        for email, marcas in contagem.items()
        if len(marcas) > 1
    ]
    duplicatas.sort(key=lambda x: -x["qtd"])
    return jsonify(duplicatas)


@app.route("/api/exportar")
def api_exportar():
    """
    Export server-side — suporta CSV e XLSX com filtros aplicados.

    Params:
        formato  : csv | xlsx  (padrão: csv)
        + todos os filtros de _aplicar_filtros_query()
    """
    fmt    = request.args.get("formato", "csv").lower()
    filtros = {k: v for k, v in request.args.items() if k != "formato"}

    snap = _snapshot_dados()
    _aplicar_url_health(snap)
    _aplicar_afiliados_health(snap)
    dados_filtrados = _aplicar_filtros_query(snap, filtros) if filtros else snap

    COLUNAS = [
        "marca", "razao_social", "cnpj", "url", "email_contato", "status",
        "url_afiliados", "status_afiliados",
        "regime_tributario", "porte_empresa", "situacao_cadastral", "capital_social",
        "natureza_juridica", "data_abertura", "logradouro", "numero", "complemento",
        "bairro", "municipio", "uf", "cep", "pais",
        "fonte_regime", "confiabilidade_dado", "data_coleta", "observacao",
    ]

    nome_arquivo = f"bets_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bets"
            # Cabeçalho estilizado
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(color="E2E8F0", bold=True)
            for col_idx, col_name in enumerate(COLUNAS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill   = header_fill
                cell.font   = header_font
                cell.alignment = Alignment(horizontal="center")
            # Dados
            for row_idx, r in enumerate(dados_filtrados, 2):
                for col_idx, col in enumerate(COLUNAS, 1):
                    ws.cell(row=row_idx, column=col_idx, value=r.get(col, ""))
            # Autofit colunas
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return Response(
                buf.read(),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}.xlsx"'},
            )
        except ImportError:
            # Fallback para CSV se openpyxl não estiver instalado
            pass

    # CSV (padrão + fallback)
    def csv_celula(v):
        s = str(v or "")
        if any(c in s for c in (',', '"', '\n')):
            return '"' + s.replace('"', '""') + '"'
        return s

    linhas = [",".join(COLUNAS)]
    for r in dados_filtrados:
        linhas.append(",".join(csv_celula(r.get(c, "")) for c in COLUNAS))

    conteudo = "﻿" + "\r\n".join(linhas)
    return Response(
        conteudo,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}.csv"'},
    )


@app.route("/api/snapshots")
def api_snapshots():
    """Retorna histórico diário de KPIs para sparklines dinâmicas."""
    try:
        import stats_snapshot
        return jsonify(stats_snapshot.ler_snapshots())
    except Exception:
        return jsonify([])


@app.route("/api/notificacoes/config", methods=["GET", "POST"])
def api_notificacoes_config():
    """GET retorna config atual; POST atualiza."""
    try:
        import notificacoes
        if request.method == "POST":
            nova = request.get_json(silent=True) or {}
            notificacoes.salvar_config(nova)
            return jsonify({"ok": True, "config": notificacoes.ler_config()})
        return jsonify(notificacoes.ler_config())
    except Exception as e:
        return jsonify({"erro": str(e)}), 500


@app.route("/api/notificacoes/teste", methods=["POST"])
def api_notificacoes_teste():
    """Dispara um webhook de teste com payload de exemplo."""
    try:
        import notificacoes
        ok, msg = notificacoes.disparar_teste()
        return jsonify({"ok": ok, "mensagem": msg})
    except Exception as e:
        return jsonify({"ok": False, "mensagem": str(e)}), 500


@app.route("/api/sistema")
def api_sistema():
    """Informações do ambiente (versões, dependências disponíveis)."""
    import sys
    deps = {}
    for lib in ("bs4", "waitress", "playwright", "openpyxl", "curl_cffi", "pandas"):
        try:
            m = __import__(lib)
            deps[lib] = getattr(m, "__version__", "ok")
        except ImportError:
            deps[lib] = None
    return jsonify({
        "python":              sys.version.split()[0],
        "flask":               __import__("flask").__version__,
        "playwright_disponivel": _PLAYWRIGHT_DISPONIVEL,
        "dependencias":        deps,
        "total_registros":     len(_dados),
        "arquivo_json_existe": ARQUIVO_JSON.exists(),
    })


# ---------------------------------------------------------------------------
# Inicialização
# ---------------------------------------------------------------------------

recarregar_dados()


def _deve_iniciar_worker() -> bool:
    if __name__ != "__main__":
        return True
    return os.environ.get("WERKZEUG_RUN_MAIN") == "true"


if _deve_iniciar_worker():
    url_health.iniciar_worker()
    csv_sync.iniciar_worker()
    if _AFILIADOS_HEALTH_DISPONIVEL:
        afiliados_health.iniciar_worker()
    if _RA_HEALTH_DISPONIVEL:
        _ra_health.iniciar_worker()


if __name__ == "__main__":
    recarregar_dados()
    fonte = "JSON enriquecido" if ARQUIVO_JSON.exists() else "CSV básico"
    n_ov  = sum(1 for r in _dados if r.get("_editado_manualmente"))
    print(f"\nDashboard iniciado — {len(_dados)} registros ({fonte})")
    if n_ov:
        print(f"  {n_ov} edição(ões) manual(is) aplicada(s)")
    print("Acesse: http://localhost:5000\n")
    app.run(debug=True, port=5000)
